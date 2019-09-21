import eel
import openpyxl
import tkinter as tk
from tkinter import filedialog
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import time

@eel.expose
def importFile():
	
	path = fileName
	wb_obj = openpyxl.load_workbook(path) 

	sheet_obj = wb_obj["Sheet1"]


	maxRow = sheet_obj.max_row

	for i in range(maxRow + 1):
		print(sheet_obj.cell(row=i+1,column=1))
	pass


@eel.expose
def openFileDialog():
	root = tk.Tk()
	root.withdraw()
	root.wm_attributes('-topmost', 1)
	file_path = filedialog.askopenfilename()

	wb_obj = openpyxl.load_workbook(file_path)

	sheet_obj = wb_obj["Sheet1"]
	global activeSheet
	activeSheet = sheet_obj 
	maxCol = sheet_obj.max_column
	maxRow = sheet_obj.max_row

	eel.createTable("varTable",maxRow,maxCol)
	print(maxRow,maxCol)

	for i in range(maxRow + 1):
		for j in range(maxCol +1):
			newCell = sheet_obj.cell(row=i+1,column=j+1)
			newCell = newCell.value
			print(newCell)
			eel.insertValue('varTable',j,newCell)
	pass

@eel.expose
def getCellValue(_rawText):
	global activeSheet

	rawText = _rawText

	maxCol = activeSheet.max_column
	maxRow = activeSheet.max_row

	for i in range(maxCol+1):
		targetVar = "@var" + str((i+1))
		targetNewline = "\n"
		varValue = activeSheet.cell(row=1,column=i+1)
		varValue = str(varValue.value)
		rawText = rawText.replace(targetVar,varValue)
		rawText = rawText.replace(targetNewline,"<br>")

	return rawText

	pass

@eel.expose
def startTask(_contactVar,msg,_delayTime):

	global activeSheet
	maxRow = activeSheet.max_row
	maxCol = activeSheet.max_column

	eel.returnTaskLog('Current Task: ' + str(maxRow) + ' contacts in total.')

	originalMsg = msg
	dynamicMsg = originalMsg

	driver = webdriver.Chrome()
	driver.get("https://web.whatsapp.com/")

	_contactVar = str(_contactVar)
	_contactVar = _contactVar.replace('var','')
	_contactVar = int(_contactVar)

	eel.returnTaskLog('Establishing Internet Connection...')

	while True:
		try:
			driver.find_element_by_xpath('/html/body/div[1]/div/div/div[3]/div/div[1]/div/label/input')
			eel.returnTaskLog('Pending Log In..')
			time.sleep(10)
			break
		except NoSuchElementException:  #spelling error making this code not work as expected
			pass

	for i in range(maxRow + 1):
		dynamicMsg = originalMsg
		phoneNum = activeSheet.cell(row=i+1,column=_contactVar)
		phoneNum = phoneNum.value
		
		for j in range(maxCol + 1):
			targetVar = "@var"+ str((j+1))
			varValue = activeSheet.cell(row=i+1,column=j+1)
			varValue = varValue.value
			dynamicMsg = dynamicMsg.replace(targetVar,str(varValue))

		if phoneNum is not None:
			driver.get('https://web.whatsapp.com/send?phone=' + str(phoneNum))
			eel.returnTaskLog( str((i+1)) +  '. Sending to target [' + str(phoneNum) + ']')
			while True:
				try:
					input_box = driver.find_element_by_xpath('/html/body/div[1]/div/div/div[4]/div/footer/div[1]/div[2]/div/div[2]')
					for ch in dynamicMsg:
					    if ch == "\n":
					        ActionChains(driver).key_down(Keys.SHIFT).key_down(Keys.ENTER).key_up(Keys.ENTER).key_up(Keys.SHIFT).key_up(Keys.BACKSPACE).perform()
					    else:
					        input_box.send_keys(ch)
					input_box.send_keys(Keys.ENTER)
					time.sleep(int(_delayTime))
					break
				except NoSuchElementException:  #spelling error making this code not work as expected
					pass

	eel.returnTaskLog('Task completed successfully.')
	pass


activeSheet = None
eel.init('web')
eel.start('indexv2.html',size=(900,1080))
